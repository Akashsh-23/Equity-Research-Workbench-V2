"""Generate formatted Excel reports for company health scores."""

import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT  = Font(name="Calibri", size=16, bold=True, color="1F2937")
LABEL_FONT  = Font(name="Calibri", size=11, bold=True)
THIN_BORDER = Border(*(Side(style="thin", color="D1D5DB"),) * 4)


def _style_header_row(ws, row, n_cols):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def build_report(ticker, info, ratios, score):
    """
    Returns an in-memory Excel file (BytesIO) with three sheets:
    Overview, Ratios, Peer Comparison.
    """
    wb = Workbook()

    ws = wb.active
    ws.title = "Overview"
    ws["A1"] = info.get("name", ticker)
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Ticker: {ticker}    Sector: {info.get('sector', 'N/A')}    " \
               f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, color="6B7280")

    ws["A4"] = "Health Score"
    ws["A4"].font = LABEL_FONT
    ws["B4"] = score["overall_score"]
    ws["B4"].font = Font(size=20, bold=True, color="2563EB")
    ws["C4"] = score["grade"]

    ws["A6"] = "Category Breakdown"
    ws["A6"].font = LABEL_FONT
    row = 7
    ws.cell(row=row, column=1, value="Category").font = HEADER_FONT
    ws.cell(row=row, column=2, value="Percentile Score").font = HEADER_FONT
    _style_header_row(ws, row, 2)
    for cat, val in score["category_scores"].items():
        row += 1
        ws.cell(row=row, column=1, value=cat).border = THIN_BORDER
        cell = ws.cell(row=row, column=2, value=val)
        cell.number_format = "0.0"
        cell.border = THIN_BORDER

    sample = score["sample_info"]
    row += 2
    ws.cell(row=row, column=1, value="Compared against").font = LABEL_FONT
    ws.cell(row=row, column=2,
            value=f"{sample['sample_size']} companies "
                  f"({'sector: ' + sample['sector'] if sample['compared_against']=='sector' else 'whole dataset'})")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 35

    ws2 = wb.create_sheet("Ratios")
    ws2.cell(row=1, column=1, value="Ratio").font = HEADER_FONT
    ws2.cell(row=1, column=2, value="Value").font = HEADER_FONT
    _style_header_row(ws2, 1, 2)
    for i, (label, value) in enumerate(ratios.items(), start=2):
        ws2.cell(row=i, column=1, value=label).border = THIN_BORDER
        cell = ws2.cell(row=i, column=2, value=value)
        cell.number_format = "0.00"
        cell.border = THIN_BORDER
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 16

    ws3 = wb.create_sheet("Peer Comparison")
    from core.database import get_all_snapshots
    all_snaps = get_all_snapshots()
    sector = info.get("sector")
    peers = [s for s in all_snaps if s.get("sector") == sector] if sector else all_snaps

    headers = ["Ticker", "Name", "Net Margin %", "ROE %", "ROCE %",
               "Current Ratio", "Debt to Equity", "Revenue Growth %"]
    for col, h in enumerate(headers, start=1):
        ws3.cell(row=1, column=col, value=h)
    _style_header_row(ws3, 1, len(headers))

    for r, peer in enumerate(peers, start=2):
        values = [
            peer.get("ticker"), peer.get("name"), peer.get("net_margin"),
            peer.get("roe"), peer.get("roce"), peer.get("current_ratio"),
            peer.get("debt_to_equity"), peer.get("revenue_growth"),
        ]
        for col, v in enumerate(values, start=1):
            cell = ws3.cell(row=r, column=col, value=v)
            cell.border = THIN_BORDER
            if peer.get("ticker") == ticker:
                cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")

    for col in range(1, len(headers) + 1):
        ws3.column_dimensions[get_column_letter(col)].width = 16

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
